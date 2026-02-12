"""
Service export dữ liệu sang định dạng AMIS
Hỗ trợ export Excel, XML, JSON cho cả Invoice và Order
"""

from typing import Optional, Dict, Any, List
from pathlib import Path
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from loguru import logger

from backend.config import settings
from backend.invoice import InvoiceData, ExportRequest


class AMISExportService:
    """Service export dữ liệu sang định dạng AMIS"""
    
    def __init__(self):
        """Khởi tạo service"""
        self.output_dir = Path(settings.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(
        self,
        invoice_data: InvoiceData,
        job_id: str,
        template_type: str = "general"
    ) -> str:
        """
        Export dữ liệu hóa đơn sang file Excel theo định dạng AMIS

        Args:
            invoice_data: Dữ liệu hóa đơn
            job_id: ID của job
            template_type: Loại template

        Returns:
            Đường dẫn đến file Excel đã tạo
        """
        try:
            logger.info(f"Bắt đầu export Excel cho job {job_id}")

            # Tạo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hóa đơn GTGT"

            # ========== STYLES ==========
            # Title style
            title_font = Font(bold=True, size=16, color="1F4E78")
            subtitle_font = Font(bold=True, size=11, color="1F4E78")

            # Header styles
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Info section styles
            label_font = Font(bold=True, size=10)
            value_font = Font(size=10)

            # Border styles
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            thick_border = Border(
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )

            # Total row styles
            total_font = Font(bold=True, size=11)
            total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            current_row = 1

            # ========== HEADER - TIÊU ĐỀ ==========
            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value="HÓA ĐƠN GIÁ TRỊ GIA TĂNG")
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value="VAT INVOICE")
            cell.font = Font(bold=True, size=10, italic=True, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2

            # ========== THÔNG TIN NGƯỜI BÁN ==========
            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value="THÔNG TIN NGƯỜI BÁN")
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            current_row += 1

            seller_info = [
                ["Tên đơn vị:", invoice_data.seller_legal_name or ""],
                ["Mã số thuế:", invoice_data.seller_tax_code or ""],
                ["Địa chỉ:", invoice_data.seller_address or ""],
            ]

            for label, value in seller_info:
                ws.cell(row=current_row, column=1, value=label).font = label_font
                ws.merge_cells(f'B{current_row}:J{current_row}')
                ws.cell(row=current_row, column=2, value=value).font = value_font
                current_row += 1

            current_row += 1

            # ========== THÔNG TIN HÓA ĐƠN ==========
            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value="THÔNG TIN HÓA ĐƠN")
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            current_row += 1

            invoice_info = [
                ["Ký hiệu hóa đơn:", invoice_data.inv_series or ""],
                ["Ngày hóa đơn:", invoice_data.inv_date.strftime("%d/%m/%Y") if invoice_data.inv_date else ""],
                ["Hình thức thanh toán:", invoice_data.payment_method_name or "TM/CK"],
            ]

            for label, value in invoice_info:
                ws.cell(row=current_row, column=1, value=label).font = label_font
                ws.merge_cells(f'B{current_row}:J{current_row}')
                ws.cell(row=current_row, column=2, value=value).font = value_font
                current_row += 1

            current_row += 1

            # ========== THÔNG TIN NGƯỜI MUA ==========
            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value="THÔNG TIN NGƯỜI MUA")
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            current_row += 1

            buyer_info = [
                ["Tên đơn vị:", invoice_data.buyer_legal_name or ""],
                ["Mã số thuế:", invoice_data.buyer_tax_code or ""],
                ["Địa chỉ:", invoice_data.buyer_address or ""],
                ["Số điện thoại:", invoice_data.buyer_phone_number or ""],
                ["Email:", invoice_data.buyer_email or ""],
            ]

            for label, value in buyer_info:
                ws.cell(row=current_row, column=1, value=label).font = label_font
                ws.merge_cells(f'B{current_row}:J{current_row}')
                ws.cell(row=current_row, column=2, value=value).font = value_font
                current_row += 1

            current_row += 2

            # ========== BẢNG CHI TIẾT HÀNG HÓA ==========
            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value="CHI TIẾT HÀNG HÓA / DỊCH VỤ")
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            current_row += 1

            # Header bảng
            headers = [
                ("STT", 5),
                ("Mã hàng", 15),
                ("Tên hàng hóa/dịch vụ", 35),
                ("ĐVT", 10),
                ("Số lượng", 10),
                ("Đơn giá", 15),
                ("Thành tiền", 15),
                ("Thuế suất", 10),
                ("Tiền thuế", 15),
                ("Tổng tiền", 15)
            ]

            for col_idx, (header, width) in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            current_row += 1
            detail_start_row = current_row

            # Dữ liệu chi tiết
            for item in invoice_data.original_invoice_detail:
                row_data = [
                    item.sort_order or item.line_number,
                    item.item_code or "",
                    item.item_name or "",
                    item.unit_name or "",
                    float(item.quantity),
                    float(item.unit_price),
                    float(item.amount_without_vat),
                    item.vat_rate_name or "10%",
                    float(item.vat_amount),
                    float(item.amount_without_vat + item.vat_amount)
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.font = value_font

                    # Alignment
                    if col_idx == 1:  # STT
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx in [3]:  # Tên hàng
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    elif col_idx in [5, 6, 7, 9, 10]:  # Các cột số tiền
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                    elif col_idx == 4:  # Số lượng
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.##'
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.row_dimensions[current_row].height = 25
                current_row += 1

            # ========== TỔNG CỘNG ==========
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws.cell(row=current_row, column=1, value="TỔNG CỘNG")
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border

            # Tổng thành tiền
            cell = ws.cell(row=current_row, column=7, value=float(invoice_data.total_amount_without_vat))
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = thick_border

            # Cột thuế suất (trống)
            cell = ws.cell(row=current_row, column=8, value="")
            cell.fill = total_fill
            cell.border = thick_border

            # Tổng tiền thuế
            cell = ws.cell(row=current_row, column=9, value=float(invoice_data.total_vat_amount))
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = thick_border

            # Tổng thanh toán
            cell = ws.cell(row=current_row, column=10, value=float(invoice_data.total_amount))
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = thick_border

            current_row += 2

            # ========== TỔNG TIỀN BẰNG CHỮ ==========
            ws.cell(row=current_row, column=1, value="Tổng tiền bằng chữ:").font = label_font
            ws.merge_cells(f'B{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=2, value=invoice_data.total_amount_in_words or "")
            cell.font = Font(size=10, italic=True)
            current_row += 2

            # ========== CHỮ KÝ ==========
            signature_row = current_row

            # Người mua
            ws.merge_cells(f'A{signature_row}:C{signature_row}')
            cell = ws.cell(row=signature_row, column=1, value="NGƯỜI MUA HÀNG")
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Người bán
            ws.merge_cells(f'H{signature_row}:J{signature_row}')
            cell = ws.cell(row=signature_row, column=8, value="NGƯỜI BÁN HÀNG")
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1
            ws.merge_cells(f'A{current_row}:C{current_row}')
            cell = ws.cell(row=current_row, column=1, value="(Ký, ghi rõ họ tên)")
            cell.font = Font(size=9, italic=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(f'H{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=8, value="(Ký, đóng dấu, ghi rõ họ tên)")
            cell.font = Font(size=9, italic=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # ========== FOOTER ==========
            current_row += 6
            ws.merge_cells(f'A{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=1, value=f"Xuất bởi AMIS OCR System - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            cell.font = Font(size=8, italic=True, color="999999")
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Lưu file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"HoaDon_GTGT_{job_id}_{timestamp}.xlsx"
            file_path = self.output_dir / filename

            wb.save(file_path)
            logger.info(f"Đã export Excel: {file_path}")

            return str(file_path)

        except Exception as e:
            logger.error(f"Lỗi khi export Excel: {str(e)}")
            raise
    
    def export_to_json(
        self,
        invoice_data: InvoiceData,
        job_id: str
    ) -> str:
        """
        Export dữ liệu sang JSON
        
        Args:
            invoice_data: Dữ liệu hóa đơn
            job_id: ID của job
            
        Returns:
            Đường dẫn đến file JSON
        """
        try:
            import json
            from decimal import Decimal
            
            # Custom JSON encoder cho Decimal và date
            class CustomEncoder(json.JSONEncoder):
                def default(self, obj):
                    if isinstance(obj, Decimal):
                        return float(obj)
                    if isinstance(obj, datetime):
                        return obj.isoformat()
                    if hasattr(obj, '__dict__'):
                        return obj.__dict__
                    return super().default(obj)
            
            # Convert sang dict
            data_dict = invoice_data.model_dump()
            
            # Lưu file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"amis_invoice_{job_id}_{timestamp}.json"
            file_path = self.output_dir / filename
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data_dict, f, ensure_ascii=False, indent=2, cls=CustomEncoder)
            
            logger.info(f"Đã export JSON: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Lỗi khi export JSON: {str(e)}")
            raise

    # ==================== ORDER EXPORT METHODS ====================

    def export_order_to_json(self, order_data, job_id: str) -> str:
        """
        Export dữ liệu đơn hàng ra file JSON

        Args:
            order_data: OrderData object
            job_id: ID của job

        Returns:
            Đường dẫn file JSON đã tạo
        """
        try:
            import json
            
            # Prepare data
            export_data = self._prepare_order_data_for_export(order_data)

            # Tạo filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"order_{job_id}_{timestamp}.json"
            file_path = self.output_dir / filename

            # Write JSON
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)

            logger.info(f"✅ Đã export order JSON: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"❌ Lỗi export order JSON: {str(e)}")
            raise

    def export_order_to_excel(self, order_data, job_id: str) -> str:
        """
        Export dữ liệu đơn hàng ra file Excel

        Args:
            order_data: OrderData object
            job_id: ID của job

        Returns:
            Đường dẫn file Excel đã tạo
        """
        try:
            # Prepare data
            export_data = self._prepare_order_data_for_export(order_data)

            # Tạo filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"order_{job_id}_{timestamp}.xlsx"
            file_path = self.output_dir / filename

            # Tạo Excel với pandas và openpyxl
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet 1: Order Information
                order_info_df = pd.DataFrame([export_data["order_info"]])
                order_info_df.to_excel(writer, sheet_name='Order Information', index=False)

                # Sheet 2: Order Items
                if export_data["items"]:
                    items_df = pd.DataFrame(export_data["items"])
                    items_df.to_excel(writer, sheet_name='Order Items', index=False)
                else:
                    # Tạo sheet rỗng với headers
                    items_df = pd.DataFrame(columns=[
                        "order_id", "item_no", "product_code", "product_name",
                        "quantity", "unit", "unit_price", "total_price", "note"
                    ])
                    items_df.to_excel(writer, sheet_name='Order Items', index=False)

                # Sheet 3: Metadata
                metadata_df = pd.DataFrame([export_data["metadata"]])
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

            # Format Excel
            self._format_order_excel_file(file_path)

            logger.info(f"✅ Đã export order Excel: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"❌ Lỗi export order Excel: {str(e)}")
            raise

    def _prepare_order_data_for_export(self, order_data) -> Dict[str, Any]:
        """
        Chuẩn bị dữ liệu đơn hàng cho export

        Args:
            order_data: OrderData object

        Returns:
            Dict với dữ liệu đã format
        """
        # Convert Decimal sang float/str cho JSON serialization
        def convert_decimal(obj):
            if isinstance(obj, Decimal):
                return float(obj) if obj else None
            return obj

        # Convert date sang string
        order_date_str = order_data.order_date.isoformat() if order_data.order_date else ""

        # Prepare order info
        order_info = {
            "order_id": order_data.order_id or "",
            "customer_id": getattr(order_data, 'customer_id', "") or "",
            "customer_name": order_data.customer_name or "",
            "customer_phone": order_data.customer_phone or "",
            "customer_address": order_data.customer_address or "",
            "order_date": order_date_str,
            "total_amount": convert_decimal(order_data.total_amount),
            "total_items": len(order_data.items),
            "status": getattr(order_data, 'status', "") or "",
            "customer_email": order_data.customer_email or "",
            "customer_type": order_data.customer_type or "",
            "business_name": order_data.business_name or "",
            "customer_tax_code": order_data.customer_tax_code or "",
            "business_address": order_data.business_address or "",
            "payment_method": order_data.payment_method or "",
            "notes": order_data.notes or "",
        }

        # Prepare items
        items = []
        for item in order_data.items:
            items.append({
                "order_id": order_data.order_id or "",
                "item_no": item.line_number,
                "product_code": getattr(item, 'product_code', "") or "",
                "product_name": item.product_name or "",
                "quantity": convert_decimal(item.quantity),
                "unit": getattr(item, 'unit', "") or "",
                "unit_price": convert_decimal(item.unit_price),
                "total_price": convert_decimal(item.total_price),
                "note": item.notes or "",
            })

        return {
            "order_info": order_info,
            "items": items,
            "metadata": {
                "processing_time": order_data.processing_time,
                "needs_review": order_data.needs_review,
                "review_notes": order_data.review_notes or "",
                "noise_detected": order_data.noise_detected or [],
                "exported_at": datetime.now().isoformat(),
            }
        }

    def _format_order_excel_file(self, file_path: Path):
        """
        Format file Excel cho đẹp

        Args:
            file_path: Đường dẫn file Excel
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)

            # Style cho headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_alignment = Alignment(horizontal="center", vertical="center")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Format headers (row 1)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment

                # Auto-adjust column width
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            logger.debug(f"✅ Đã format Excel file: {file_path}")

        except Exception as e:
            logger.warning(f"⚠️ Không thể format Excel: {str(e)}")


# Singleton instance
amis_export_service = AMISExportService()